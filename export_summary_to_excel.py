# export_summary_to_excel.py
# ============================================================
# Export ASR_Summary.txt results to a formatted Excel workbook
# ============================================================
#
# Reads the final_redteam_results.csv and ASR_Summary.txt and
# creates a professional multi-sheet Excel report with all data.
#
# Usage:
#   python export_summary_to_excel.py
#
# Output:
#   results/ASR_Research_Report.xlsx (multi-sheet workbook)
# ============================================================

import pandas as pd
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Run: pip install openpyxl")


def export_summary_to_excel(
    csv_path="results/final_redteam_results.csv",
    txt_path="results/ASR_Summary.txt",
    output_path="results/ASR_Research_Report.xlsx"
):
    """
    Creates a professionally formatted Excel report from the results.
    
    Sheets:
    1. Overall_Summary    — Key metrics (ASR, Refusal Rate, model info)
    2. Strategy_Breakdown — Per-strategy ASR with status indicators
    3. Risk_Category      — Per-risk-category ASR breakdown
    4. Successful_Attacks — Details of attacks that achieved unsafe compliance
    5. All_Results        — Full item-level data
    6. Text_Report        — The complete ASR_Summary.txt content
    """
    
    if not os.path.exists(csv_path):
        print(f"ERROR: {csv_path} not found. Run the pipeline first.")
        return
    
    df = pd.read_csv(csv_path)
    
    if df.empty:
        print("ERROR: Results CSV is empty.")
        return
    
    # ─── Calculate all metrics ───
    total = len(df)
    unsafe_count = (df["unsafe_compliance"] == True).sum() if "unsafe_compliance" in df.columns else 0
    refused_count = (df["refused"] == True).sum() if "refused" in df.columns else 0
    asr = round((unsafe_count / total * 100), 2) if total > 0 else 0
    refusal_rate = round((refused_count / total * 100), 2) if total > 0 else 0
    ambiguous = total - unsafe_count - refused_count
    
    red_model = df["red_model"].iloc[0] if "red_model" in df.columns else "unknown"
    blue_model = df["blue_model"].iloc[0] if "blue_model" in df.columns else "unknown"
    judge_model = df["judge_model"].iloc[0] if "judge_model" in df.columns else "unknown"
    
    # ─── Sheet 1: Overall Summary ───
    overall_data = pd.DataFrame([
        {"Metric": "Report Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Metric": "", "Value": ""},
        {"Metric": "Red Agent Model", "Value": str(red_model)},
        {"Metric": "Blue Agent Model", "Value": str(blue_model)},
        {"Metric": "Judge Model", "Value": str(judge_model)},
        {"Metric": "", "Value": ""},
        {"Metric": "Total Attack Prompts", "Value": total},
        {"Metric": "Unsafe Compliance (Attacks Won)", "Value": unsafe_count},
        {"Metric": "Blue Refused", "Value": refused_count},
        {"Metric": "Neither (Ambiguous)", "Value": ambiguous},
        {"Metric": "", "Value": ""},
        {"Metric": "ASR (Attack Success Rate) %", "Value": f"{asr}%"},
        {"Metric": "Refusal Rate %", "Value": f"{refusal_rate}%"},
        {"Metric": "", "Value": ""},
        {"Metric": "Winner", "Value": "RED TEAM" if asr >= 50 else "BLUE TEAM"},
        {"Metric": "Assessment", "Value": 
            "HIGH ASR - Red Team dominates" if asr >= 75 else
            "MODERATE ASR - Competitive" if asr >= 50 else
            "LOW ASR - Blue Team advantage" if asr >= 25 else
            "VERY LOW ASR - Red attacks mostly failed"
        },
    ])
    
    # ─── Sheet 2: Strategy Breakdown ───
    strategy_rows = []
    if "strategy" in df.columns:
        for strategy in sorted(df["strategy"].unique()):
            mask = df["strategy"] == strategy
            s_total = mask.sum()
            s_unsafe = (df.loc[mask, "unsafe_compliance"] == True).sum()
            s_refused = (df.loc[mask, "refused"] == True).sum() if "refused" in df.columns else 0
            s_asr = round((s_unsafe / s_total * 100), 2) if s_total > 0 else 0
            
            if s_asr >= 50: status = "HIGH"
            elif s_asr >= 25: status = "MEDIUM"
            elif s_asr > 0: status = "LOW"
            else: status = "ZERO"
            
            strategy_rows.append({
                "Strategy": strategy,
                "Total_Attacks": s_total,
                "Unsafe_Compliance": s_unsafe,
                "Refused": s_refused,
                "Ambiguous": s_total - s_unsafe - s_refused,
                "ASR_%": s_asr,
                "Status": status
            })
    
    strategy_df = pd.DataFrame(strategy_rows)
    
    # ─── Sheet 3: Risk Category Breakdown ───
    risk_rows = []
    if "risk_category" in df.columns:
        for cat in sorted(df["risk_category"].dropna().unique()):
            mask = df["risk_category"] == cat
            c_total = mask.sum()
            c_unsafe = (df.loc[mask, "unsafe_compliance"] == True).sum()
            c_asr = round((c_unsafe / c_total * 100), 2) if c_total > 0 else 0
            risk_rows.append({
                "Risk_Category": str(cat),
                "Total_Attacks": c_total,
                "Unsafe_Compliance": c_unsafe,
                "ASR_%": c_asr
            })
    
    risk_df = pd.DataFrame(risk_rows)
    if not risk_df.empty:
        risk_df = risk_df.sort_values("ASR_%", ascending=False)
    
    # ─── Sheet 4: Successful Attacks ───
    success_df = pd.DataFrame()
    if "unsafe_compliance" in df.columns:
        success_mask = df["unsafe_compliance"] == True
        if success_mask.any():
            cols = ["task_id", "seed_instruction", "risk_category", "strategy", 
                    "test_prompt", "blue_response", "judge_notes"]
            available_cols = [c for c in cols if c in df.columns]
            success_df = df.loc[success_mask, available_cols].copy()
            # Truncate long text for readability
            for col in ["test_prompt", "blue_response", "seed_instruction"]:
                if col in success_df.columns:
                    success_df[col] = success_df[col].astype(str).str[:300]
    
    # ─── Sheet 5: All Results (truncated for Excel size) ───
    all_results = df.copy()
    for col in ["test_prompt", "blue_response", "seed_instruction", "judge_notes"]:
        if col in all_results.columns:
            all_results[col] = all_results[col].astype(str).str[:200]
    
    # ─── Sheet 6: Text Report ───
    text_report_rows = []
    if os.path.exists(txt_path):
        with open(txt_path, "r", encoding="utf-8") as f:
            for line in f:
                text_report_rows.append({"Report_Line": line.rstrip()})
    text_report_df = pd.DataFrame(text_report_rows)
    
    # ═══════════════════════════════════════════
    # Write to Excel with formatting
    # ═══════════════════════════════════════════
    
    if HAS_OPENPYXL:
        wb = Workbook()
        
        # ── Style definitions ──
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        title_font = Font(name="Calibri", size=16, bold=True, color="2F5496")
        metric_font = Font(name="Calibri", size=11, bold=True)
        value_font = Font(name="Calibri", size=11)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        def write_dataframe_to_sheet(ws, df_data, start_row=1):
            """Write a DataFrame to a worksheet with formatting."""
            # Headers
            for col_idx, col_name in enumerate(df_data.columns, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Data rows
            for row_idx, row in enumerate(df_data.itertuples(index=False), start_row + 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = value_font
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Auto-width columns
            for col_idx in range(1, len(df_data.columns) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(df_data.columns[col_idx - 1])),
                    df_data.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df_data) > 0 else 0
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        
        # ── Sheet 1: Overall Summary ──
        ws1 = wb.active
        ws1.title = "Overall_Summary"
        ws1.cell(row=1, column=1, value="RED TEAM vs BLUE TEAM — ASR RESEARCH REPORT").font = title_font
        ws1.merge_cells("A1:B1")
        write_dataframe_to_sheet(ws1, overall_data, start_row=3)
        
        # Highlight ASR row
        for row in ws1.iter_rows(min_row=3, max_row=20, min_col=1, max_col=2):
            for cell in row:
                if cell.value and "ASR" in str(cell.value):
                    cell.font = Font(name="Calibri", size=14, bold=True, color="FF0000")
                if cell.value and "Winner" in str(cell.value):
                    cell.font = Font(name="Calibri", size=14, bold=True)
        
        # ── Sheet 2: Strategy Breakdown ──
        ws2 = wb.create_sheet("Strategy_Breakdown")
        ws2.cell(row=1, column=1, value="ASR BREAKDOWN BY STRATEGY").font = title_font
        if not strategy_df.empty:
            write_dataframe_to_sheet(ws2, strategy_df, start_row=3)
            # Color-code ASR cells
            for row_idx in range(4, 4 + len(strategy_df)):
                asr_cell = ws2.cell(row=row_idx, column=6)  # ASR_% column
                status_cell = ws2.cell(row=row_idx, column=7)  # Status column
                try:
                    val = float(asr_cell.value) if asr_cell.value else 0
                    if val >= 50:
                        asr_cell.fill = green_fill
                        status_cell.fill = green_fill
                    elif val > 0:
                        asr_cell.fill = yellow_fill
                        status_cell.fill = yellow_fill
                    else:
                        asr_cell.fill = red_fill
                        status_cell.fill = red_fill
                except (ValueError, TypeError):
                    pass
        
        # ── Sheet 3: Risk Category ──
        ws3 = wb.create_sheet("Risk_Category")
        ws3.cell(row=1, column=1, value="ASR BREAKDOWN BY RISK CATEGORY").font = title_font
        if not risk_df.empty:
            write_dataframe_to_sheet(ws3, risk_df, start_row=3)
        
        # ── Sheet 4: Successful Attacks ──
        ws4 = wb.create_sheet("Successful_Attacks")
        ws4.cell(row=1, column=1, value=f"SUCCESSFUL ATTACKS ({len(success_df)} total)").font = title_font
        if not success_df.empty:
            write_dataframe_to_sheet(ws4, success_df, start_row=3)
        else:
            ws4.cell(row=3, column=1, value="No successful attacks recorded.")
        
        # ── Sheet 5: All Results ──
        ws5 = wb.create_sheet("All_Results")
        ws5.cell(row=1, column=1, value="FULL ITEM-LEVEL RESULTS").font = title_font
        write_dataframe_to_sheet(ws5, all_results, start_row=3)
        
        # ── Sheet 6: Text Report ──
        ws6 = wb.create_sheet("Text_Report")
        ws6.cell(row=1, column=1, value="COMPLETE TEXT REPORT (from ASR_Summary.txt)").font = title_font
        if not text_report_df.empty:
            for row_idx, row in enumerate(text_report_df.itertuples(index=False), 3):
                cell = ws6.cell(row=row_idx, column=1, value=row[0])
                cell.font = Font(name="Consolas", size=10)
            ws6.column_dimensions["A"].width = 100
        
        # Save
        wb.save(output_path)
        print(f"✅ Excel report saved to: {output_path}")
        print(f"   Sheets: Overall_Summary, Strategy_Breakdown, Risk_Category, Successful_Attacks, All_Results, Text_Report")
    
    else:
        # Fallback: use pandas ExcelWriter without formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overall_data.to_excel(writer, sheet_name='Overall_Summary', index=False)
            if not strategy_df.empty:
                strategy_df.to_excel(writer, sheet_name='Strategy_Breakdown', index=False)
            if not risk_df.empty:
                risk_df.to_excel(writer, sheet_name='Risk_Category', index=False)
            if not success_df.empty:
                success_df.to_excel(writer, sheet_name='Successful_Attacks', index=False)
            all_results.to_excel(writer, sheet_name='All_Results', index=False)
            if not text_report_df.empty:
                text_report_df.to_excel(writer, sheet_name='Text_Report', index=False)
        
        print(f"✅ Excel report saved to: {output_path}")
    
    # Also print a quick summary
    print(f"\n📊 Quick Summary:")
    print(f"   Total Attacks:      {total}")
    print(f"   Unsafe Compliance:  {unsafe_count}")
    print(f"   ASR:                {asr}%")
    print(f"   Winner:             {'RED TEAM' if asr >= 50 else 'BLUE TEAM'}")


if __name__ == "__main__":
    export_summary_to_excel()
