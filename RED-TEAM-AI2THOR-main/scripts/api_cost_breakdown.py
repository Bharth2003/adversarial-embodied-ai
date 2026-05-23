from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Sheet 1: Per-Turn Token Estimates ---
ws1 = wb.active
ws1.title = "Token Estimates"

header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
header_fill = PatternFill("solid", fgColor="2D2D2D")
section_fill = PatternFill("solid", fgColor="3A1F5E")
data_font = Font(name="Arial", size=10)
number_font = Font(name="Arial", size=10, color="0000FF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Token breakdown per evaluation
ws1["A1"] = "TOKEN BREAKDOWN PER EVALUATION CALL"
ws1["A1"].font = Font(bold=True, size=13, name="Arial", color="FFFFFF")
ws1["A1"].fill = section_fill
ws1.merge_cells("A1:C1")

headers = ["Component", "Est. Tokens", "Notes"]
for i, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border

data = [
    ["Execution report (steps + outcomes)", 500, "~15-20 steps with state changes"],
    ["Safety constitution / rules", 200, "Constitution rules fed to judge"],
    ["Instruction + ground truth label", 100, "Original task + safe/unsafe label"],
    ["Judge prompt template", 150, "System prompt for evaluation"],
    ["TOTAL INPUT per eval", None, ""],
    ["Judge output (JSON verdict + reason)", 200, "Structured JSON response"],
    ["TOTAL OUTPUT per eval", None, ""],
]

for r, row in enumerate(data, 3):
    for c_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if c_idx == 2 and isinstance(val, int):
            cell.font = number_font

ws1["B7"] = "=SUM(B3:B6)"
ws1["B7"].font = Font(bold=True, name="Arial", size=10)
ws1["B9"] = "=B8"
ws1["B9"].font = Font(bold=True, name="Arial", size=10)

for row in [7, 9]:
    for col in range(1, 4):
        ws1.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E8E0F0")

ws1.column_dimensions["A"].width = 42
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 38

# --- Sheet 2: Cost Table ---
ws2 = wb.create_sheet("Cost Comparison")

ws2["A1"] = "CLAUDE API COST COMPARISON - EVALUATION JUDGE"
ws2["A1"].font = Font(bold=True, size=13, name="Arial", color="FFFFFF")
ws2["A1"].fill = section_fill
ws2.merge_cells("A1:H1")

# Assumptions
ws2["A3"] = "ASSUMPTIONS"
ws2["A3"].font = Font(bold=True, size=11, name="Arial")
ws2["B4"] = "Input tokens per eval:"
ws2["C4"] = 950
ws2["B5"] = "Output tokens per eval:"
ws2["C5"] = 200
ws2["B6"] = "Turns per person:"
ws2["C6"] = 300
for r in range(4, 7):
    ws2.cell(row=r, column=2).font = data_font
    ws2.cell(row=r, column=3).font = number_font

# Pricing row
ws2["A8"] = "MODEL PRICING (per million tokens)"
ws2["A8"].font = Font(bold=True, size=11, name="Arial")

price_headers = ["Model", "Input $/1M", "Output $/1M", "Quality", "Speed", "Best For"]
for i, h in enumerate(price_headers, 1):
    c = ws2.cell(row=9, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border

prices = [
    ["Claude Haiku 4.5", 0.80, 4.00, "Good", "Very Fast", "Budget eval - recommended"],
    ["Claude Sonnet 4.6", 3.00, 15.00, "Very Good", "Fast", "High accuracy eval"],
    ["Claude Opus 4.6", 15.00, 75.00, "Best", "Slower", "Overkill for eval"],
    ["GPT-4o-mini", 0.15, 0.60, "Good", "Fast", "Cheapest option"],
    ["GPT-4o", 2.50, 10.00, "Very Good", "Fast", "Alternative to Sonnet"],
    ["Qwen 2.5 (local)", 0, 0, "Decent", "Depends", "Free - your Ollama setup"],
]

for r, row in enumerate(prices, 10):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if c_idx in [2, 3] and isinstance(val, (int, float)):
            cell.font = number_font
            cell.number_format = '$#,##0.00'

# Cost scenarios
ws2["A18"] = "COST PER SCENARIO"
ws2["A18"].font = Font(bold=True, size=11, name="Arial")

scenario_headers = ["Model", "1 Person\n300 turns", "3 People\n900 turns", "5 People\n1500 turns", "8 People\n2400 turns", "8 People x3 runs\n7200 turns"]
for i, h in enumerate(scenario_headers, 1):
    c = ws2.cell(row=19, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = thin_border

# Row 20: Haiku
models_rows = [
    ["Claude Haiku 4.5", 10],
    ["Claude Sonnet 4.6", 11],
    ["Claude Opus 4.6", 12],
    ["GPT-4o-mini", 13],
    ["GPT-4o", 14],
    ["Qwen 2.5 (local)", 15],
]

turn_counts = [300, 900, 1500, 2400, 7200]

for r, (model, price_row) in enumerate(models_rows, 20):
    ws2.cell(row=r, column=1, value=model).font = Font(bold=True, name="Arial", size=10)
    ws2.cell(row=r, column=1).border = thin_border

    for col_idx, turns in enumerate(turn_counts, 2):
        # Formula: (turns * input_tokens / 1000000 * input_price) + (turns * output_tokens / 1000000 * output_price)
        # input_price = B{price_row}, output_price = C{price_row}
        # input_tokens = C4, output_tokens = C5
        formula = f"=({turns}*$C$4/1000000*B${price_row})+({turns}*$C$5/1000000*C${price_row})"
        cell = ws2.cell(row=r, column=col_idx, value=formula)
        cell.number_format = '$#,##0.00'
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# Highlight free row
for col in range(1, 7):
    ws2.cell(row=25, column=col).fill = PatternFill("solid", fgColor="D4EDDA")

# Highlight recommended
for col in range(1, 7):
    ws2.cell(row=20, column=col).fill = PatternFill("solid", fgColor="E8E0F0")

ws2.column_dimensions["A"].width = 22
for col in range(2, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# --- Sheet 3: Recommendation ---
ws3 = wb.create_sheet("Recommendation")

ws3["A1"] = "RECOMMENDATION FOR YOUR GROUP"
ws3["A1"].font = Font(bold=True, size=13, name="Arial", color="FFFFFF")
ws3["A1"].fill = section_fill
ws3.merge_cells("A1:B1")

recs = [
    ["Approach", "Details"],
    ["Primary Judge", "Qwen 2.5 on your Ollama (free, everyone has access)"],
    ["Validation Judge", "Claude Haiku API - one key, run once to validate"],
    ["Est. validation cost", "~$1-3 for entire project"],
    ["", ""],
    ["Why Hybrid?", ""],
    ["1. Reproducibility", "Local model = same results every time, no API dependency"],
    ["2. Cost", "Qwen is free for unlimited runs during development"],
    ["3. Accuracy check", "Run Claude Haiku once at the end to validate Qwen's judgments"],
    ["4. Paper-friendly", "You can report agreement rate between local and API judge"],
    ["", ""],
    ["Setup", ""],
    ["Task model", "dolphin3:8b or qwen2.5 (your Ollama)"],
    ["Safety model", "paulprt1:latest (your Ollama)"],
    ["Judge model (local)", "qwen2.5-32k:latest (your Ollama)"],
    ["Judge model (validation)", "Claude Haiku via API ($0.80/$4.00 per M tokens)"],
]

for r, row in enumerate(recs, 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if r == 2:
            cell.font = header_font
            cell.fill = header_fill

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 60

output_path = os.path.join(os.path.dirname(__file__), "..", "data", "api_cost_breakdown.xlsx")
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Saved to {output_path}")
